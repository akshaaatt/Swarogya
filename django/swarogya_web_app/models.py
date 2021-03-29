from django.core.files import File
from django.core.files.storage import FileSystemStorage,Storage
from json_excel_converter import Converter
from json_excel_converter.xlsx import Writer
import re
import firebase_admin
from firebase_admin import credentials
from firebase_admin import auth
from firebase_admin import firestore
from firebase_admin import storage
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import _datetime
import datetime
import os
from pathlib import Path
import pyrebase
import json
import requests
BASE_DIR = Path(__file__).resolve(strict=True).parent.parent
# Imports the Google Cloud client library
from google.cloud import storage

# Instantiates a client

with open('config_file.json') as json_file:
  data = json.load(json_file)

firebase_app = pyrebase.initialize_app(data)
authentic = firebase_app.auth()

cred = credentials.Certificate('firebase-sdk.json')
firebase_admin.initialize_app(cred,{
    'storageBucket': 'swarogya-aa30d.appspot.com'
})
db = firestore.client()

client = storage.Client()

bucket = client.get_bucket('swarogya-aa30d.appspot.com')

storage = firebase_app.storage()

def get_employees(hospitalName):
    meta = list()
    doc_ref = db.collection('Employees').where('hospitalName','==',hospitalName).stream()
    for docs in doc_ref:
        meta.append(docs.to_dict())
    return meta

def removeEmployee(email):
    staff_user = auth.get_user_by_email(email)
    auth.delete_user(staff_user.uid)
    doc_ref = db.collection(u'Employees').where(u'emailId', u'==', str(email)).stream()
    for docs in doc_ref:
        db.collection(u'Employees').document(docs.id).delete()
    return "Health Professional Removed "


def get_patient_report(my_doc):
    add_color_report([{'Patient Name' :my_doc['name'],'Patient ID':my_doc['patientId'], 'Analysis' :sorted(my_doc['measurements'],reverse=False)}])

def get_hospital_name(emailId):
    meta = {}
    doc_ref = db.collection('Employees').where('emailId','==',emailId).stream()
    for docs in doc_ref:
        meta['hospitalName'] = docs.get('hospitalName')
        meta['admin'] = docs.get('admin')
        meta['doc_id'] = docs.id
    print(meta)
    return meta

def check_if_patientId_present(patientId):
    try:
        doc_doc = db.collection('OngoingTreatments').where('patientId', '==', patientId).stream()
        for docs in doc_ref:
            for doc,value in vars(docs).items():
                if doc == '_data':
                    value.pop("recordsLastUpdatedOn", None)
                    value.pop("joinedOn", None)
                    meta = value
        return "ID Present"
    except NameError as err:
        return "ID not Present"

def get_pic_url(pic,pic_name1,pic_name2,patientId, uid):

    storage.child("OngoingTreatments/"+patientId+"/Documents/"+pic_name1+"/"+pic_name2).put(pic)
    return storage.child("OngoingTreatments/"+patientId+"/Documents/"+pic_name1+"/"+pic_name2).get_url(auth.create_custom_token(uid))

def get_week_range(my_week):
    weekdates = list()
    week = my_week.replace("W", '').split('-')
    for day in range(7):
        week_date = datetime.datetime.strptime('{}-W{}'.format(week[0], int(week[1]) - 1) + '-{}'.format(day),
                                               "%Y-W%W-%w")
        weekdates.append(week_date.strftime("%Y-%m-%d"))
    return weekdates

dates = {}

def loop_for_dates(items, value,analysis_parameter):
    my_list = [[], [], [], [], []]

    if analysis_parameter == 'Date-Wise':
        for key,loop in value.items():
            dates.update({key:loop})
    else:
        for key,loop in value.items():

            my_list[0].append(loop['BP-Low'])
            my_list[1].append(loop['Pulse'])
            my_list[2].append(loop['BP-High'])
            my_list[3].append(loop['Oxygen'])
            my_list[4].append(loop['Temperature'])

        dates.update({items: {'BP-Low': sum(my_list[0]) / len(my_list[0]),
                                                    'BP-High': sum(my_list[2]) / len(my_list[2]),
                                                    'Pulse': sum(my_list[1]) / len(my_list[1]),
                                                    'Oxygen': sum(my_list[3]) / len(my_list[3]),
                                                    'Temperature': sum(my_list[4]) / len(my_list[4])}})


def analysis_func(doc, analysis_parameter, analysis_parameter_value, type_analysis):
    patient_list = list()
    check = "no"
    for docs in doc:
        try:
            if docs.get('measurements'):
                for items, values in sorted(docs.get('measurements').items(), reverse=False):
                    if analysis_parameter == 'Month-Wise':
                        if analysis_parameter_value in items:
                            loop_for_dates(items, values,'Month-Wise')
                            check = "yes"
                    elif analysis_parameter == 'Week-Wise':
                        week = get_week_range(analysis_parameter_value)
                        if any(items in s for s in week):
                            loop_for_dates(items, values,'Week-Wise')
                            check = "yes"
                    else:
                        if items == analysis_parameter_value:
                            loop_for_dates(items, values,'Date-Wise')
                            check = "yes"
                val = dates.copy()
                if check == 'yes':
                    patient_list.append({'Patient Name' : docs.get('name'), 'Patient ID':docs.get('patientId'),'Analysis':{ docs.get('hospitalName')+' - '+docs.get('treatmentMode')+' - '+type_analysis+' - '+analysis_parameter + " - " + analysis_parameter_value : val}})
                    check ="no"
            dates.clear()
        except KeyError as er:
            continue
    if patient_list:
        add_color_report(patient_list)
        return (analysis_parameter +" Reported Generated for " + analysis_parameter_value + " at " + str( _datetime.date.today()) )
    else:
        return "No " + analysis_parameter + " Records available in given time frame to Generate Report "

def get_context(db,date_val, week_val, month_val, type_analysis,hospitalName):
    context = {}
    context['hospitalName'] = hospitalName
    if date_val != None:
        context['to_display'] = analysis_func(db, 'Date-Wise', date_val, type_analysis)
    elif week_val != None:
        context['to_display'] = analysis_func(db, 'Week-Wise', week_val, type_analysis)
    else:
        context['to_display'] = analysis_func(db, 'Month-Wise', month_val, type_analysis)
    return context

class Employee:
    def __init__(self, name, staff_id, phone_no, email, designation, hospitalName,admin='no'):
        self.name = name
        self.staffId = staff_id
        self.phoneNumber = phone_no
        self.emailId = email
        self.designation = designation
        self.hospitalName = hospitalName
        self.admin = admin

    def add_doc_firebase(self):
        doc_ref = db.collection('Employees').add(vars(self))

def user_auth(email,password):
    user = authentic.create_user_with_email_and_password(email,password)
    staff_user = auth.get_user_by_email(email)
    return staff_user.uid

class ManualPatient:
    def __init__(self, name, address, contact_number,dob,gender, bloodGroup,email_id, first_emergency_contact,first_emergency_contact_number,
    second_emergency_contact, second_emergency_contact_number, barcode, hospitalName,  wingNumber, floorNumber, roomNumber, bedNumber,treatmentMode,
    face_pic = None,documents='None', joined_on=firestore.SERVER_TIMESTAMP,status = "active" ):
        self.name = name
        self.address = address
        self.phoneNumber = contact_number
        self.gender = gender
        self.dob = dob
        self.bloodGroup = bloodGroup
        self.emailId = email_id
        self.firstEmergencyContactName = first_emergency_contact
        self.firstEmergencyContactNumber = first_emergency_contact_number
        self.secondEmergencyContactName = second_emergency_contact
        self.secondEmergencyContactNumber = second_emergency_contact_number
        self.patientId = barcode
        self.hospitalName = hospitalName
        self.wingNumber = wingNumber
        self.floorNumber = floorNumber
        self.roomNumber = roomNumber
        self.bedNumber = bedNumber
        self.joinedOn = joined_on
        self.facePic = face_pic
        self.documents = documents
        self.joinedOn = joined_on
        self.treatmentMode = treatmentMode
        self.status = status

    def add_patient_firebase(self):
        doc = db.collection('OngoingTreatments').where('hospitalName', '==', self.hospitalName).where('wingNumber', '==', self.wingNumber).where('floorNumber', '==', self.floorNumber).where('roomNumber', '==', self.roomNumber).where('bedNumber', '==', self.bedNumber).stream()
        try:
            for docs in doc:
                meta = docs.to_dict()
            if meta:
                return "Bed already occupied"
        except UnboundLocalError as err:
            doc_ref = db.collection('OngoingTreatments').add(vars(self))
            return "Patient Added"


class UpdatePatient:
    def __init__(self, name, address, contact_number,gender,dob, bloodGroup,email_id,first_emergency_contact,first_emergency_contact_number,
    second_emergency_contact, second_emergency_contact_number, barcode, hospitalName,  wingNumber, floorNumber, roomNumber, bedNumber):
        self.name = name
        self.address = address
        self.phoneNumber = contact_number
        self.gender = gender
        self.dob = dob
        self.bloodGroup = bloodGroup
        self.emailId = email_id
        self.firstEmergencyContactName = first_emergency_contact
        self.firstEmergencyContactNumber = first_emergency_contact_number
        self.secondEmergencyContactName = second_emergency_contact
        self.secondEmergencyContactNumber = second_emergency_contact_number
        self.patientId = barcode
        self.hospitalName = hospitalName
        self.wingNumber = wingNumber
        self.floorNumber = floorNumber
        self.roomNumber = roomNumber
        self.bedNumber = bedNumber
        self.treatmentMode = treatmentMode

def treated_patient(patient_barcode, status):

    doc_ref = db.collection('OngoingTreatments').where('barcode', '==', patient_barcode).stream()
    try:
        for doc in doc_ref:
            my_doc_id = doc.id
        db.collection('OngoingTreatments').document(my_doc_id).update({u'status': status})
        db.collection('OngoingTreatments').document(my_doc_id).delete()
        return "Patient Removed"
    except UnboundLocalError as error:
        return "Patient ID not present"

class SmartPatient:
    def __init__(self,barcode, hospitalName,  wingNumber, floorNumber, roomNumber, bedNumber,documents='None',status = "active",treatmentMode="Hospitalized"):
        self.patientId = barcode
        self.hospitalName = hospitalName
        self.wingNumber = wingNumber
        self.floorNumber = floorNumber
        self.roomNumber = roomNumber
        self.bedNumber = bedNumber
        self.documents = documents
        self.status = status
        self.treatmentMode = treatmentMode

    def add_patient_firebase(self):
        doc = db.collection('OngoingTreatments').where('hospitalName', '==', self.hospitalName).where('wingNumber', '==', self.wingNumber).where('floorNumber', '==', self.floorNumber).where('roomNumber', '==', self.roomNumber).where('bedNumber', '==', self.bedNumber).stream()
        try:
            for docs in doc:
                meta = docs.to_dict()
            if meta:
                return "Bed already occupied"
        except UnboundLocalError as err:
            doc_ref = db.collection('OngoingTreatments').add(vars(self))
            return "Patient Added"

class QuarantinePatient:
    def __init__(self,barcode, hospitalName,documents='None',status = "active",treatmentMode='Home Care'):
        self.patientId = barcode
        self.hospitalName = hospitalName
        self.documents = documents
        self.status = status
        self.treatmentMode = treatmentMode

    def add_patient_firebase(self):
        doc_ref = db.collection('OngoingTreatments').add(vars(self))
        return "Patient Added"

class Patient:
    def __init__(self, name, address, contact_number, email_id, emergency_contact,
    emergency_contact_number ):
        self.name = name
        self.address = address
        self.contact_number = contact_number
        self.email_id = email_id
        self.emergency_contact = emergency_contact
        self.emergency_contact_number = emergency_contact_number
        self.joined_on = joined_on

def wing_analysis_db(hospitalName, wingNumber):
    context = {}
    context['doc'] = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('wingNumber', '==', wingNumber).stream()
    context['type_analysis'] ="Hospital Name:" + hospitalName + " - Wing Number:"+wingNumber
    return context

def hospital_analysis_db(hospitalName):
    context = {}
    context['doc'] = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('treatmentMode', '==', 'Hospitalized').stream()
    context['type_analysis'] ="Hospital Name:" + hospitalName
    return context

def home_quarantine_db(hospitalName):
    context = {}
    context['doc'] = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('treatmentMode', '==', 'Home Care').stream()
    context['type_analysis'] ="Hospital Name:" + hospitalName
    return context

def floor_analysis_db(hospitalName, wingNumber, floorNumber):
    context = {}
    context['doc'] = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('wingNumber', '==', wingNumber).where('floorNumber', '==', floorNumber).stream()
    context['type_analysis'] ="Hospital Name:" + hospitalName + " - Wing Number:"+wingNumber + ' - Floor Number:' +  floorNumber
    return context

def room_analysis_db(hospitalName, wingNumber, floorNumber, roomNumber):
    context = {}
    context['doc'] = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('wingNumber', '==', wingNumber).where('floorNumber', '==', floorNumber).where('roomNumber', '==', roomNumber).stream()
    context['type_analysis'] ="Hospital Name:" + hospitalName + " - Wing Number:"+wingNumber + ' - Floor Number:' +  floorNumber + ' - Room Number:' +  roomNumber
    return context


def get_searchPatient(patientId,hospitalName):
    doc_ref = db.collection('OngoingTreatments').where('hospitalName', '==', hospitalName).where('patientId', '==', patientId).stream()
    try:
        for docs in doc_ref:
            for doc,value in vars(docs).items():
                if doc == '_data':
                    value.pop("recordsLastUpdatedOn", None)
                    value.pop("joinedOn", None)
                    meta = value
            id = docs.id

        return meta,id
    except NameError as err:
        return "No Registered Patient with the Hospital"

def update_patient_details(doc_id,patient):
    doc_ref = db.collection('OngoingTreatments').document(doc_id)
    doc_ref.update(vars(patient))

def add_color_report(doc):
    conv = Converter()
    conv.convert(doc, Writer(file=os.path.join(BASE_DIR,"media/report.xlsx")))
    workbook = load_workbook(filename=os.path.join(BASE_DIR,"media/report.xlsx"))
    sheet = workbook.active
    red_background = PatternFill(bgColor="FF0000")
    diff_style = DifferentialStyle(fill=red_background)

    for row in sheet.iter_rows(min_row=4,
                               max_row=4,
                               min_col=2):
        import re
        for cell in row:
            if cell.value != None:
                res = re.findall(r'(\w+?)(\d+)', cell.coordinate)[0]
                if cell.value == 'Temperature':
                    rulee = CellIsRule(">", ("98.8",), stopIfTrue=False, fill=red_background, )
                    sheet.conditional_formatting.add("{}5:{}1048576".format(res[0], res[0]), rulee)
                elif cell.value == 'BP-Low':
                    rulee = CellIsRule(">", ("80",), stopIfTrue=False, fill=red_background, )
                    sheet.conditional_formatting.add("{}5:{}1048576".format(res[0], res[0]), rulee)
                elif cell.value == 'BP-High':
                    rulee = CellIsRule(">", ("120",), stopIfTrue=False, fill=red_background, )
                    sheet.conditional_formatting.add("{}5:{}1048576".format(res[0], res[0]), rulee)
                elif cell.value == 'Pulse':
                    rulee = CellIsRule(operator='between', formula=['50', '100'], stopIfTrue=True, fill=red_background)
                    sheet.conditional_formatting.add("{}5:{}1048576".format(res[0], res[0]), rulee)
                elif cell.value == 'Oxygen':
                    sheet.conditional_formatting.add("{}5:{}1048576".format(res[0], res[0]) ,CellIsRule(operator='between', formula=['1', '95'], stopIfTrue=True, fill=red_background))

    workbook.save(filename=os.path.join(BASE_DIR,"media/report.xlsx"))
